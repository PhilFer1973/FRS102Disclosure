"""Assemble the issues register (pipeline stage 9 output): an Excel workbook of
findings with a full audit trail, plus the question round and a summary.

Data-only (no formulas), professional styling, reviewer-disposition dropdowns.
Findings: numerical (casts, cross-casts, ratios, OCR-suspect, low-confidence) and
disclosure (required-but-missing, present-unverified). Every row carries its
citation and the reasoning, so the reviewer can trace each finding to source.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Calibri"
HEADER_FILL = PatternFill("solid", start_color="1F3864")
MISSING_FILL = PatternFill("solid", start_color="F8CBAD")     # required & missing
NUMERIC_FILL = PatternFill("solid", start_color="FCE4D6")

ISSUE_HEADERS = ["#", "Category", "Citation", "Severity", "Finding",
                 "Evidence / source", "Status", "Reviewer disposition"]
ISSUE_WIDTHS = [5, 22, 14, 26, 70, 50, 12, 26]
Q_HEADERS = ["Fact key", "Question", "Affects", "Answer (true/false/value)"]
Q_WIDTHS = [40, 70, 30, 26]

SEVERITY_LABEL = {"statutory": "Statutory (materiality-blind)",
                  "standard-material": "Standard - material",
                  "standard-immaterial-candidate": "Standard - immaterial candidate"}


def _header(ws, headers, widths):
    for col, (h, w) in enumerate(zip(headers, widths, strict=True), start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def _issue_rows(numerical, presence) -> list[dict]:
    rows: list[dict] = []
    for p in presence:
        if p.status == "present":
            continue
        req = p.requirement.requirement
        rows.append({
            "category": "Disclosure - MISSING" if p.status == "absent"
                        else "Disclosure - verify present",
            "citation": f"{req.source} {req.reference}",
            "severity": SEVERITY_LABEL.get(req.severity, req.severity),
            "finding": (("Required disclosure not found in the accounts: "
                         if p.status == "absent"
                         else "Could not confirm this required disclosure - verify: ")
                        + req.requirement_text),
            "evidence": p.evidence, "status": "open", "fill": MISSING_FILL})
    for f in numerical:
        rows.append({
            "category": "Numerical - " + f.check_type,
            "citation": f.location,
            "severity": SEVERITY_LABEL.get(f.severity, f.severity),
            "finding": f.description,
            "evidence": "" if not f.is_error else "could not evaluate - confirm",
            "status": "unverified" if f.is_error else "open", "fill": NUMERIC_FILL})
    return rows


def build_register(out_path: str | Path, entity: str, period_end: str, edition: str,
                   numerical: list, presence: list, questions: list,
                   counts: dict | None = None) -> Path:
    wb = Workbook()
    body = Font(name=FONT, size=10)
    wrap = Alignment(vertical="top", wrap_text=True)

    # Summary
    ws = wb.active
    ws.title = "Summary"
    info = [("FRS 102 disclosure review - issues register", True),
            ("", False), (f"Entity: {entity}", False),
            (f"Period end: {period_end}", False),
            (f"FRS 102 edition: {edition}", False), ("", False)]
    miss = sum(1 for p in presence if p.status == "absent")
    verify = sum(1 for p in presence if p.status == "unclear")
    num_real = sum(1 for f in numerical if not f.is_error)
    num_err = sum(1 for f in numerical if f.is_error)
    info += [(f"Required disclosures missing: {miss}", False),
             (f"Disclosures to verify present: {verify}", False),
             (f"Numerical findings: {num_real} ({num_err} could not be evaluated)", False),
             (f"Open questions for the reviewer: {len(questions)}", False)]
    if counts:
        info += [("", False), ("Checklist outcomes: " + ", ".join(
            f"{k}={v}" for k, v in counts.items()), False)]
    info += [("", False),
             ("Every finding is a candidate for the reviewer's professional "
              "judgement; this register is not a signed opinion.", False)]
    for n, (text, bold) in enumerate(info, start=1):
        ws.cell(row=n, column=1, value=text).font = Font(
            name=FONT, bold=bold, size=13 if bold else 10)
    ws.column_dimensions["A"].width = 95

    # Issues
    ws = wb.create_sheet("Issues register")
    _header(ws, ISSUE_HEADERS, ISSUE_WIDTHS)
    rows = _issue_rows(numerical, presence)
    for r, item in enumerate(rows, start=2):
        values = [r - 1, item["category"], item["citation"], item["severity"],
                  item["finding"], item["evidence"], item["status"], ""]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = body
            cell.alignment = wrap
        ws.cell(row=r, column=2).fill = item["fill"]
    n = len(rows) + 1
    dv = DataValidation(type="list", allow_blank=True,
                        formula1='"accept,reject,amend,immaterial"')
    ws.add_data_validation(dv)
    dv.add(f"H2:H{max(n, 2)}")
    ws.auto_filter.ref = f"A1:H{max(n, 2)}"

    # Questions
    ws = wb.create_sheet("Questions")
    _header(ws, Q_HEADERS, Q_WIDTHS)
    for r, q in enumerate(questions, start=2):
        for c, v in enumerate([q.fact_key, q.question_text,
                               ", ".join(q.affected_refs), ""], start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = body
            cell.alignment = wrap

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
