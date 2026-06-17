from openpyxl import load_workbook

from pipeline.assemble.register import build_register
from pipeline.engine.checklist import EngineResult, Requirement
from pipeline.engine.presence import PresenceResult
from pipeline.engine.questions import Question
from pipeline.validate.checks import Finding


def _req(ref):
    return Requirement("id", "FRS102", ref, "both", "Disclose X", "always", None,
                       (), "missing", "standard-material")


def test_build_register(tmp_path):
    numerical = [
        Finding("cast", "balance_sheet:total", "does not cast: ...",
                "standard-material"),
        Finding("low_confidence", "balance_sheet:share_capital",
                "low OCR confidence", "standard-material", is_error=True),
    ]
    presence = [
        PresenceResult(EngineResult(_req("7.3"), "applicable"), "absent",
                       "no cash flow statement found"),
        PresenceResult(EngineResult(_req("4.12"), "applicable"), "present",
                       "see note 22"),
        PresenceResult(EngineResult(_req("3.9"), "applicable"), "unclear", ""),
    ]
    questions = [Question("is_consolidated", "Are these consolidated accounts?",
                          ("9.23",))]
    out = build_register(tmp_path / "reg.xlsx", "Demo Ltd", "2024-12-31",
                         "pre-PR2024", numerical, presence, questions)
    wb = load_workbook(out)
    assert wb.sheetnames == ["Summary", "Issues register", "Questions"]
    issues = wb["Issues register"]
    # 2 numerical + missing + unclear = 4 issue rows (present one is excluded)
    assert issues.max_row - 1 == 4
    cats = [issues.cell(r, 2).value for r in range(2, issues.max_row + 1)]
    assert any("MISSING" in c for c in cats)
    assert wb["Questions"].cell(2, 1).value == "is_consolidated"
