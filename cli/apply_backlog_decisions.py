"""Apply reviewer decisions on the in_review backlog to the requirements table.

Two input modes:
  --from-xlsx review/backlog_review.xlsx   (default) — read the reviewer's
      'YOUR DECISION' column (ACTIVATE / REJECT / HOLD); blank falls back to
      Claude's recommendation in that row.
  --recommendations                        — apply build/backlog_recommendations
      .json directly (activate -> activate, reject -> rejected, review -> HOLD).

Safeguards:
  - De-dup on activate: an exact-duplicate rule (same citation/edition/direction/
    trigger/text) is activated only once; the rest stay in_review, so duplicate
    findings are not re-introduced.
  - Activation guard: any unregistered trigger fact on an activating row is
    registered first (inferred value_type), so the DB guard cannot abort.
  - --dry-run prints the plan and writes nothing.

Activation is the reviewer's decision; this tool only executes it.

  uv run python -m cli.apply_backlog_decisions --dry-run
  uv run python -m cli.apply_backlog_decisions --recommendations
  uv run python -m cli.apply_backlog_decisions --from-xlsx review/backlog_review.xlsx
"""

from __future__ import annotations

import argparse
import json
import os
from pathlib import Path

import psycopg
from dotenv import load_dotenv
from openpyxl import load_workbook

from db.load_requirements import infer_value_type

RECS_JSON = Path("build/backlog_recommendations.json")
DEFAULT_XLSX = Path("review/backlog_review.xlsx")
_REC_TO_DECISION = {"activate": "ACTIVATE", "reject": "REJECT", "review": "HOLD"}


def _dsn() -> str:
    load_dotenv()
    dsn = os.environ.get("SUPABASE_DB_URL")
    if not dsn:
        raise SystemExit("SUPABASE_DB_URL not set.")
    return dsn


def decisions_from_recs() -> dict[str, str]:
    recs = json.loads(RECS_JSON.read_text(encoding="utf-8"))
    return {r["id"]: _REC_TO_DECISION[r["recommend"]] for r in recs}


def decisions_from_xlsx(path: Path) -> dict[str, str]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Backlog"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rec_col = headers.index("Recommendation") + 1
    dec_col = headers.index("YOUR DECISION") + 1
    id_col = headers.index("ID (do not edit)") + 1
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        rid = ws.cell(r, id_col).value
        if not rid:
            continue
        decision = (ws.cell(r, dec_col).value or "").strip().upper()
        if decision not in ("ACTIVATE", "REJECT", "HOLD"):
            decision = _REC_TO_DECISION.get(
                (ws.cell(r, rec_col).value or "").strip().lower(), "HOLD")
        out[str(rid)] = decision
    return out


def _identity(row: dict) -> tuple:
    return (row["source"], row["reference"], row["edition"], row["direction"],
            row["trigger_type"], row["trigger_condition"] or "",
            row["requirement_text"])


def apply(decisions: dict[str, str], dry_run: bool) -> None:
    with psycopg.connect(_dsn(), autocommit=False,
                         prepare_threshold=None) as conn, conn.cursor() as cur:
        # current in_review rows + their identity, and identities already active
        cur.execute("select id, source, reference, edition, requirement_text, "
                    "trigger_type, trigger_condition, trigger_facts, direction "
                    "from requirements where status = 'in_review'")
        cols = ("id", "source", "reference", "edition", "requirement_text",
                "trigger_type", "trigger_condition", "trigger_facts", "direction")
        backlog = {str(r[0]): dict(zip(cols, r, strict=True)) for r in cur.fetchall()}
        cur.execute("select source, reference, edition, direction, trigger_type, "
                    "coalesce(trigger_condition,''), requirement_text "
                    "from requirements where status = 'active'")
        active_identities = {tuple(r) for r in cur.fetchall()}
        cur.execute("select key from fact_registry")
        registered = {r[0] for r in cur.fetchall()}

        to_activate, to_reject, dup_skipped, new_facts = [], [], [], set()
        seen_identity = set(active_identities)
        for rid, decision in decisions.items():
            row = backlog.get(rid)
            if row is None:
                continue  # already actioned / not in backlog
            if decision == "REJECT":
                to_reject.append(rid)
            elif decision == "ACTIVATE":
                ident = _identity(row)
                if ident in seen_identity:
                    dup_skipped.append(rid)
                    continue
                seen_identity.add(ident)
                to_activate.append(rid)
                new_facts.update(f for f in (row["trigger_facts"] or ())
                                 if f not in registered)
            # HOLD -> leave in_review

        print(f"backlog in_review: {len(backlog)}")
        print(f"  -> activate: {len(to_activate)}  (dup-skipped: {len(dup_skipped)})")
        print(f"  -> reject:   {len(to_reject)}")
        print(f"  -> hold:     {len(backlog) - len(to_activate) - len(to_reject) - len(dup_skipped)}")
        if new_facts:
            print(f"  registering {len(new_facts)} previously-unregistered fact(s)")
        if dry_run:
            print("\ndry run — nothing written.")
            return

        if new_facts:
            cur.executemany(
                "insert into fact_registry (key, description, value_type) "
                "values (%s,%s,%s) on conflict (key) do nothing",
                [(k, f"{k.replace('_', ' ')} (auto; review)", infer_value_type(k))
                 for k in sorted(new_facts)])
        if to_activate:
            cur.execute("update requirements set status='active' where id = any(%s)",
                        (to_activate,))
        if to_reject:
            cur.execute("update requirements set status='rejected' where id = any(%s)",
                        (to_reject,))
        conn.commit()
        cur.execute("select status, count(*) from requirements group by status "
                    "order by status")
        print("\nrequirements now:", dict(cur.fetchall()))


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    src = ap.add_mutually_exclusive_group()
    src.add_argument("--from-xlsx", nargs="?", const=str(DEFAULT_XLSX))
    src.add_argument("--recommendations", action="store_true",
                     help="apply Claude's recommendations from the JSON directly")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if args.recommendations:
        decisions = decisions_from_recs()
    else:
        path = Path(args.from_xlsx) if args.from_xlsx else DEFAULT_XLSX
        if not path.exists():
            raise SystemExit(f"{path} not found — run cli.assess_backlog first, or "
                             "use --recommendations.")
        decisions = decisions_from_xlsx(path)
    apply(decisions, args.dry_run)


if __name__ == "__main__":
    main()
