"""Assess the in_review requirements backlog into a batched sign-off workbook.

Reads every status='in_review' requirement, asks Haiku to classify each (its
nature + a recommended disposition + a one-line reason), applies conservative
safety overrides, then writes:
  - review/backlog_review.xlsx  — grouped by recommendation, with a disposition
    dropdown, so the reviewer approves in BATCHES rather than row by row.
  - build/backlog_recommendations.json — machine-readable, consumed by
    cli/apply_backlog_decisions.py.

NOTHING is activated here. Activation is human sign-off only (migration 0003:
"Claude Code drafts, never activates").

  uv run python -m cli.assess_backlog
"""

from __future__ import annotations

import json
import os
from pathlib import Path

import psycopg
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from pipeline.llm_client import LLMClient

RECS_JSON = Path("build/backlog_recommendations.json")
OUT_XLSX = Path("review/backlog_review.xlsx")
BATCH = 25

SYSTEM = """\
You are a senior UK chartered accountant triaging draft FRS 102 disclosure-
checklist rules that are held for review before activation. The checklist engine
checks whether a REQUIRED DISCLOSURE is present in a set of accounts; activated
rules fire findings/questions on a real engagement.

For each rule classify its NATURE:
- required_disclosure: a disclosure the accounts MUST contain when the trigger
  holds. This is the only nature suited to the presence checklist.
- permissive_option: states the entity MAY do / is permitted / need not do
  something (an option or an exemption), not a required disclosure.
- encouraged: an explicitly encouraged (not required) disclosure.
- recognition_measurement: a recognition or measurement rule (how to account for
  something), not a disclosure — handled elsewhere, not by a presence check.
- process_note: procedural/scope/boilerplate text imposing no testable
  disclosure (e.g. 'refer to the application guidance').

Then RECOMMEND a disposition:
- activate: only for a genuine required_disclosure that a presence check can test.
- reject: permissive_option / encouraged / recognition_measurement / process_note
  — these do not belong in a disclosure-presence checklist.
- review: genuinely borderline, or a 'present-when-not-required' direction where
  flagging an over-disclosure is risky (a policy immaterial this year may be
  material in the prior year shown, so extra disclosure is often correct).

Give a concise one-line reason. Return one assessment per rule, echoing its n.
"""

SCHEMA = {
    "type": "object",
    "properties": {
        "assessments": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "n": {"type": "integer"},
                    "nature": {"type": "string", "enum": [
                        "required_disclosure", "permissive_option", "encouraged",
                        "recognition_measurement", "process_note"]},
                    "recommend": {"type": "string",
                                  "enum": ["activate", "reject", "review"]},
                    "reason": {"type": "string"},
                },
                "required": ["n", "nature", "recommend", "reason"],
                "additionalProperties": False,
            },
        }
    },
    "required": ["assessments"],
    "additionalProperties": False,
}


def _dsn() -> str:
    load_dotenv()
    dsn = os.environ.get("SUPABASE_DB_URL")
    if not dsn:
        raise SystemExit("SUPABASE_DB_URL not set.")
    return dsn


def load_backlog() -> list[dict]:
    with psycopg.connect(_dsn(), prepare_threshold=None) as conn, conn.cursor() as cur:
        cur.execute(
            "select id, source, reference, edition, requirement_text, trigger_type, "
            "trigger_condition, trigger_facts, direction, severity "
            "from requirements where status = 'in_review' "
            "order by reference")
        cols = ("id", "source", "reference", "edition", "requirement_text",
                "trigger_type", "trigger_condition", "trigger_facts", "direction",
                "severity")
        return [dict(zip(cols, r, strict=True)) for r in cur.fetchall()]


def _override(rec: dict, row: dict) -> dict:
    """Conservative safety net on top of the model's call."""
    nature, recommend = rec["nature"], rec["recommend"]
    # never activate non-disclosure natures into a presence checklist
    if nature in ("permissive_option", "encouraged", "recognition_measurement",
                  "process_note") and recommend == "activate":
        recommend = "reject"
    # a 'present-when-not-required' nag is never auto-activated
    if row["direction"] == "untriggered" and recommend == "activate":
        recommend = "review"
    # encouraged disclosures should never be flagged as missing
    if row["trigger_type"] == "encouraged" and recommend == "activate":
        recommend = "review"
    return {**rec, "recommend": recommend}


def assess(rows: list[dict], client: LLMClient) -> list[dict]:
    out: dict[int, dict] = {}
    for start in range(0, len(rows), BATCH):
        chunk = rows[start:start + BATCH]
        listing = "\n".join(
            f"[n={start + i}] {r['source']} {r['reference']} | dir={r['direction']} "
            f"| sev={r['severity']} | trigger={r['trigger_type']} "
            f"cond={r['trigger_condition'] or '-'}\n  {r['requirement_text']}"
            for i, r in enumerate(chunk))
        res = client.complete_json("classify", SYSTEM,
                                   f"Classify and recommend for these rules:\n{listing}",
                                   SCHEMA, max_tokens=4000)
        for a in res["assessments"]:
            out[a["n"]] = a
        print(f"  assessed {min(start + BATCH, len(rows))}/{len(rows)}")
    recs = []
    for i, row in enumerate(rows):
        a = out.get(i, {"nature": "process_note", "recommend": "review",
                        "reason": "no assessment returned — review manually"})
        a = _override(a, row)
        recs.append({**row, "id": str(row["id"]),
                     "trigger_facts": list(row["trigger_facts"] or ()),
                     "nature": a["nature"], "recommend": a["recommend"],
                     "reason": a["reason"]})
    return recs


# ---- workbook ---------------------------------------------------------------
FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F3864")
FILL = {"activate": PatternFill("solid", start_color="E2EFDA"),
        "reject": PatternFill("solid", start_color="F2F2F2"),
        "review": PatternFill("solid", start_color="FFF2CC")}
HEADERS = ["Recommendation", "Citation", "Direction", "Severity", "Nature",
           "Why (Claude)", "Requirement text", "Trigger condition",
           "YOUR DECISION", "YOUR COMMENTS", "ID (do not edit)"]
WIDTHS = [15, 16, 12, 22, 20, 48, 60, 34, 15, 35, 38]
ORDER = {"activate": 0, "review": 1, "reject": 2}


def _header(ws):
    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS, strict=True), start=1):
        c = ws.cell(1, col, h)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def write_workbook(recs: list[dict]) -> None:
    recs = sorted(recs, key=lambda r: (ORDER[r["recommend"]], r["reference"]))
    counts = {k: sum(1 for r in recs if r["recommend"] == k)
              for k in ("activate", "review", "reject")}
    wb = Workbook()
    body = Font(name=FONT, size=10)
    wrap = Alignment(vertical="top", wrap_text=True)

    ws = wb.active
    ws.title = "Summary"
    lines = [
        ("FRS 102 rules backlog — sign-off", True), ("", False),
        (f"{len(recs)} rules held 'in_review'. Claude's recommendation:", False),
        (f"  ACTIVATE — genuine required disclosures: {counts['activate']}", False),
        (f"  REVIEW   — borderline / over-disclosure risk: {counts['review']}", False),
        (f"  REJECT   — not a presence-checkable disclosure: {counts['reject']}", False),
        ("", False),
        ("How to sign off (fast path):", True),
        ("  1. Accept the ACTIVATE batch wholesale, or scan it and set YOUR", False),
        ("     DECISION to REJECT/HOLD on any you disagree with.", False),
        ("  2. The REVIEW batch is where your judgement matters most.", False),
        ("  3. REJECT batch: skim; flip any you think should stay.", False),
        ("", False),
        ("YOUR DECISION options: ACTIVATE / REJECT / HOLD (blank = take Claude's", False),
        ("recommendation). Nothing changes until you run the apply step.", False),
    ]
    for n, (t, b) in enumerate(lines, start=1):
        ws.cell(n, 1, t).font = Font(name=FONT, bold=b, size=13 if b else 10)
    ws.column_dimensions["A"].width = 95

    ws = wb.create_sheet("Backlog")
    _header(ws)
    for r, rec in enumerate(recs, start=2):
        values = [rec["recommend"].upper(), f"{rec['source']} {rec['reference']}",
                  rec["direction"], rec["severity"], rec["nature"], rec["reason"],
                  rec["requirement_text"], rec["trigger_condition"] or "", "", "",
                  rec["id"]]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(r, c, v)
            cell.font = body
            cell.alignment = wrap
        ws.cell(r, 1).fill = FILL[rec["recommend"]]
    n = len(recs) + 1
    dv = DataValidation(type="list", formula1='"ACTIVATE,REJECT,HOLD"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"I2:I{n}")
    ws.auto_filter.ref = f"A1:K{n}"

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    RECS_JSON.parent.mkdir(parents=True, exist_ok=True)
    RECS_JSON.write_text(json.dumps(recs, indent=2), encoding="utf-8")
    print(f"\n{OUT_XLSX}: {len(recs)} rules "
          f"(activate {counts['activate']}, review {counts['review']}, "
          f"reject {counts['reject']})")
    print(f"{RECS_JSON}: recommendations written")


def main() -> None:
    import argparse
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--rebuild", action="store_true",
                    help="rebuild the workbook from the saved recommendations JSON "
                    "(no LLM calls)")
    args = ap.parse_args()
    if args.rebuild:
        recs = json.loads(RECS_JSON.read_text(encoding="utf-8"))
        write_workbook(recs)
        return
    rows = load_backlog()
    print(f"loaded {len(rows)} in_review rules; assessing with Haiku...")
    client = LLMClient()
    recs = assess(rows, client)
    write_workbook(recs)
    print("\n" + client.usage_summary())


if __name__ == "__main__":
    main()
