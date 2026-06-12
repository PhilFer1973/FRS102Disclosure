"""Build the human review workbook from draft rows + challenge results.

Output: review/checklist_review.xlsx
- 'Checklist rows': all draft rows, risk-sorted (challenge-disputed first, then
  statutory, then untriggered/both direction, then the rest), with YOUR VERDICT
  (OK/AMEND/REJECT dropdown) and YOUR COMMENTS columns.
- 'Fact keys': consolidated register of proposed keys with usage, plus an
  action dropdown (KEEP/RENAME/MERGE/DROP).
- 'Instructions': how to review.

No formulas — data only, so no recalculation step is required.
"""

from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F3864")
DISPUTED_FILL = PatternFill("solid", start_color="FCE4E4")
STATUTORY_FILL = PatternFill("solid", start_color="FFF2CC")

ROW_HEADERS = ["Row ID", "Section", "Reference", "Edition", "Challenge verdict",
               "Challenge issues", "Requirement text", "Trigger type",
               "Trigger condition", "Trigger facts", "Direction", "Severity",
               "Drafter notes", "YOUR VERDICT", "YOUR COMMENTS"]
ROW_WIDTHS = [10, 9, 11, 12, 13, 45, 60, 11, 32, 28, 11, 18, 45, 13, 35]

KEY_HEADERS = ["Fact key", "Rows using it", "Sections", "References",
               "YOUR ACTION", "Rename to / merge into", "YOUR COMMENTS"]
KEY_WIDTHS = [42, 13, 18, 40, 12, 28, 35]

INSTRUCTIONS = [
    ("FRS 102 checklist — draft row review", True),
    ("", False),
    ("Sheet 'Checklist rows' — one row per draft checklist requirement (545).", False),
    ("Rows are sorted by review priority (Challenge verdict column):", False),
    ("  1. material — the challenge found an issue that would change behaviour", False),
    ("     on a real engagement (wrong trigger/direction/severity/requirement).", False),
    ("  2. statutory rows not in tier 1 (materiality-blind if activated).", False),
    ("  3. minor — wording-level challenge issues only; plus rows with direction", False),
    ("     'untriggered' or 'both' (the confirm-immaterial side).", False),
    ("  4. clean rows — challenge found nothing.", False),
    ("", False),
    ("For each row set YOUR VERDICT:", False),
    ("  OK      — row is right as drafted.", False),
    ("  AMEND   — right idea, needs edits: put the edit in YOUR COMMENTS", False),
    ("            (or edit the cells directly — both are picked up).", False),
    ("  REJECT  — row should not exist (not a requirement / out of scope).", False),
    ("", False),
    ("Sheet 'Fact keys' — the proposed fact registry. For each key set", False),
    ("KEEP / RENAME / MERGE / DROP; for RENAME or MERGE fill the next column.", False),
    ("", False),
    ("Sheet 'Dropped rows' (if present) — rows the auto-amend pass proposes to", False),
    ("remove because the cited paragraph imposes no such requirement. Confirm", False),
    ("each with CONFIRM DROP or REINSTATE; nothing is deleted without you.", False),
    ("", False),
    ("Nothing becomes active from this workbook — verdicts are ingested back", False),
    ("into the draft rows, and activation only happens after your sign-off.", False),
]


def _section_sort_key(section: str) -> tuple[int, str]:
    digits = "".join(c for c in section if c.isdigit())
    return (int(digits) if digits else 99, section)


def _ref_sort_key(reference: str) -> list[tuple[int, str]]:
    key = []
    for part in reference.split("."):
        m = re.match(r"(\d+)(.*)", part)
        key.append((int(m.group(1)), m.group(2)) if m else (0, part))
    return key


def load_data(rows_dir: str = "build") -> tuple[list[dict], dict[str, dict]]:
    rows = []
    for path in sorted(Path(rows_dir).glob("section_*_rows.jsonl")):
        section = path.stem.removeprefix("section_").removesuffix("_rows")
        with path.open(encoding="utf-8") as fh:
            for n, line in enumerate(fh):
                if line.strip():
                    row = json.loads(line)
                    row.setdefault("_section", section)
                    row.setdefault("_row_id", f"{section}-{n:03d}")
                    rows.append(row)
    challenges: dict[str, dict] = {}
    ch_path = Path("build/challenge_results.jsonl")
    if ch_path.exists():
        with ch_path.open(encoding="utf-8") as fh:
            for line in fh:
                if line.strip():
                    c = json.loads(line)
                    challenges[c["row_id"]] = c
    return rows, challenges


ASPECT_WEIGHT = {"trigger": 5, "direction": 5, "severity": 3,
                 "faithfulness": 2, "fact_keys": 1}


def risk_rank(row: dict, challenge: dict | None) -> tuple[int, int]:
    """(tier, -score): lower sorts first. Within the material tier, rows whose
    material issues hit behaviour-critical aspects (trigger/direction) lead."""
    ch = challenge or {}
    grade = ch.get("grade", "material")
    score = sum(ASPECT_WEIGHT.get(i["aspect"], 1)
                for i in ch.get("issues", [])
                if i.get("materiality", "material") == "material")
    if row["severity"] == "statutory":
        score += 2
    if grade == "material":
        return (0, -score)
    if row["severity"] == "statutory":
        return (1, -score)
    if grade == "minor" or row["direction"] in ("untriggered", "both"):
        return (2, -score)
    return (3, -score)


def _style_header(ws, headers: list[str], widths: list[int]) -> None:
    for col, (head, width) in enumerate(zip(headers, widths, strict=True), start=1):
        cell = ws.cell(row=1, column=col, value=head)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28


def _base_id(row_id: str) -> str:
    return re.sub(r"-s\d+$", "", row_id)


def main() -> None:
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("--rows-dir", default="build",
                    help="build for v1 rows, build/rows_v2 for amended rows")
    ap.add_argument("--out", default="review/checklist_review.xlsx")
    args = ap.parse_args()
    v2 = "rows_v2" in args.rows_dir

    rows, challenges = load_data(args.rows_dir)
    rows.sort(key=lambda r: (risk_rank(r, challenges.get(_base_id(r["_row_id"]))),
                             _section_sort_key(r["_section"]),
                             _ref_sort_key(r["reference"])))

    wb = Workbook()

    # --- Instructions ---------------------------------------------------------
    ws = wb.active
    ws.title = "Instructions"
    for n, (text, bold) in enumerate(INSTRUCTIONS, start=1):
        cell = ws.cell(row=n, column=1, value=text)
        cell.font = Font(name=FONT, bold=bold, size=12 if bold else 10)
    ws.column_dimensions["A"].width = 90

    # --- Checklist rows -------------------------------------------------------
    ws = wb.create_sheet("Checklist rows")
    headers = list(ROW_HEADERS)
    if v2:
        headers[5] = "Changes made (auto-amend)"
    _style_header(ws, headers, ROW_WIDTHS)
    body_font = Font(name=FONT, size=10)
    wrap = Alignment(vertical="top", wrap_text=True)
    for r, row in enumerate(rows, start=2):
        ch = challenges.get(_base_id(row["_row_id"]), {})
        if v2:
            issues = row.get("change_summary", "") if row.get("amended") \
                else "(unchanged — challenge found nothing material)"
        else:
            issues = "; ".join(
                f"[{i['aspect']}/{i.get('materiality', '?')}] {i['problem']} "
                f"-> {i['suggestion']}"
                for i in ch.get("issues", []))
        values = [row["_row_id"], row["_section"], row["reference"], row["edition"],
                  ch.get("grade", ch.get("verdict", "n/a")), issues,
                  row["requirement_text"],
                  row["trigger_type"], row["trigger_condition"] or "",
                  ", ".join(row["trigger_facts"]), row["direction"], row["severity"],
                  row["review_notes"], "", ""]
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = body_font
            cell.alignment = wrap
        if ch.get("grade") == "material":
            ws.cell(row=r, column=5).fill = DISPUTED_FILL
        if row["severity"] == "statutory":
            ws.cell(row=r, column=12).fill = STATUTORY_FILL

    n_rows = len(rows) + 1
    verdict_dv = DataValidation(type="list", formula1='"OK,AMEND,REJECT"',
                                allow_blank=True, showDropDown=False)
    ws.add_data_validation(verdict_dv)
    verdict_dv.add(f"N2:N{n_rows}")
    ws.auto_filter.ref = f"A1:O{n_rows}"
    ws.freeze_panes = "D2"

    # --- Fact keys ------------------------------------------------------------
    usage: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        for key in row["trigger_facts"]:
            usage[key].append(row)
    merge_target: dict[str, str] = {}
    proposals_path = Path("build/key_merge_proposals.json")
    if proposals_path.exists():
        proposals = json.loads(proposals_path.read_text(encoding="utf-8"))
        for group in proposals["groups"]:
            for member in group["members"]:
                if member != group["canonical"]:
                    merge_target[member] = group["canonical"]
    ws = wb.create_sheet("Fact keys")
    _style_header(ws, KEY_HEADERS, KEY_WIDTHS)
    ordered = sorted(usage.items(), key=lambda kv: (-len(kv[1]), kv[0]))
    for r, (key, used_by) in enumerate(ordered, start=2):
        sections = sorted({u["_section"] for u in used_by}, key=_section_sort_key)
        refs = sorted({u["reference"] for u in used_by}, key=_ref_sort_key)
        values = [key, len(used_by), ", ".join(sections), ", ".join(refs), "",
                  merge_target.get(key, ""), ""]
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = body_font
            cell.alignment = wrap
    n_keys = len(ordered) + 1
    action_dv = DataValidation(type="list", formula1='"KEEP,RENAME,MERGE,DROP"',
                               allow_blank=True, showDropDown=False)
    ws.add_data_validation(action_dv)
    action_dv.add(f"E2:E{n_keys}")
    ws.auto_filter.ref = f"A1:G{n_keys}"
    ws.freeze_panes = "B2"

    # --- Dropped rows (v2 only) -----------------------------------------------
    dropped_path = Path(args.rows_dir) / "dropped.jsonl"
    n_dropped = 0
    if dropped_path.exists():
        ws = wb.create_sheet("Dropped rows")
        headers = ["Row ID", "Section", "Reference", "Edition",
                   "Original requirement text", "Drop reason",
                   "YOUR VERDICT", "YOUR COMMENTS"]
        _style_header(ws, headers, [10, 9, 11, 12, 60, 55, 16, 35])
        with dropped_path.open(encoding="utf-8") as fh:
            dropped_rows = [json.loads(line) for line in fh if line.strip()]
        for r, d in enumerate(dropped_rows, start=2):
            values = [d["_row_id"], d["_section"], d["reference"], d["edition"],
                      d["requirement_text"], d["drop_reason"], "", ""]
            for c, value in enumerate(values, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.font = body_font
                cell.alignment = wrap
        n_dropped = len(dropped_rows)
        drop_dv = DataValidation(type="list", formula1='"CONFIRM DROP,REINSTATE"',
                                 allow_blank=True, showDropDown=False)
        ws.add_data_validation(drop_dv)
        drop_dv.add(f"G2:G{n_dropped + 1}")
        ws.auto_filter.ref = f"A1:H{n_dropped + 1}"
        ws.freeze_panes = "B2"

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    amended = sum(1 for r in rows if r.get("amended"))
    print(f"{out}: {len(rows)} rows ({amended} amended, {n_dropped} dropped for "
          f"confirmation), {len(ordered)} fact keys")


if __name__ == "__main__":
    main()
